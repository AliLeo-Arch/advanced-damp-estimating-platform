"""Export estimates to CSV and Excel."""

from __future__ import annotations

import csv
import io
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.estimate_service import serialize_estimate
from app.models import Estimate
from app.pricing_engine import round_money
from app.quotation import build_quotation


def _safe_filename(reference: str, customer: str, suffix: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", customer.strip() or "Customer").strip("-")
    return f"{reference}-{slug or 'Customer'}-{suffix}"


def _quotation_rows(db: Session, estimate: Estimate) -> tuple[dict, list[dict]]:
    quote = build_quotation(db, estimate)
    header = {
        "reference": quote.estimate.reference,
        "customer_name": quote.estimate.customer_name,
        "company_name": quote.estimate.company_name,
        "site_address": quote.estimate.site_address,
        "postcode": quote.estimate.postcode,
        "surveyor": quote.estimate.surveyor,
        "survey_date": quote.estimate.survey_date,
        "status": quote.estimate.status,
        "issue_date": quote.issue_date or "",
        "valid_until": quote.valid_until or "",
        "payment_terms": quote.payment_terms,
        "subtotal_ex_vat": quote.estimate.sell_price,
        "vat_rate": quote.vat_rate,
        "vat_amount": quote.vat_amount,
        "total_inc_vat": quote.total_inc_vat,
    }
    lines = [
        {
            "label": line["label"],
            "description": line["description"],
            "amount": line["amount"],
        }
        for line in quote.scope_lines
    ]
    return header, lines


def render_estimate_csv(db: Session, estimate: Estimate) -> tuple[bytes, str]:
    header, lines = _quotation_rows(db, estimate)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Advanced Damp Ltd — Customer quotation export"])
    writer.writerow(["Reference", header["reference"]])
    writer.writerow(["Customer", header["customer_name"]])
    writer.writerow(["Site", f"{header['site_address']} {header['postcode']}".strip()])
    writer.writerow(["Surveyor", header["surveyor"]])
    writer.writerow(["Survey date", header["survey_date"]])
    writer.writerow(["Status", header["status"]])
    writer.writerow([])
    writer.writerow(["Work type", "Description", "Amount (GBP)"])
    for line in lines:
        writer.writerow([line["label"], line["description"], f"{line['amount']:.2f}"])
    writer.writerow([])
    writer.writerow(["Subtotal (ex VAT)", f"{header['subtotal_ex_vat']:.2f}"])
    writer.writerow(["VAT", f"{header['vat_amount']:.2f}"])
    writer.writerow(["Total (inc VAT)", f"{header['total_inc_vat']:.2f}"])
    writer.writerow([])
    writer.writerow(["Payment terms", header["payment_terms"]])
    filename = _safe_filename(header["reference"], header["customer_name"], "Quotation.csv")
    return buffer.getvalue().encode("utf-8-sig"), filename


def render_estimate_xlsx(db: Session, estimate: Estimate) -> tuple[bytes, str]:
    header, lines = _quotation_rows(db, estimate)
    data = serialize_estimate(estimate)
    wb = Workbook()

    quote_sheet = wb.active
    quote_sheet.title = "Quotation"
    bold = Font(bold=True)
    quote_sheet["A1"] = "Advanced Damp Ltd — Quotation"
    quote_sheet["A1"].font = bold
    quote_sheet.append(["Reference", header["reference"]])
    quote_sheet.append(["Customer", header["customer_name"]])
    quote_sheet.append(["Site", f"{header['site_address']} {header['postcode']}".strip()])
    quote_sheet.append(["Surveyor", header["surveyor"]])
    quote_sheet.append(["Survey date", header["survey_date"]])
    quote_sheet.append(["Status", header["status"]])
    quote_sheet.append([])
    quote_sheet.append(["Work type", "Description", "Amount (GBP)"])
    for cell in quote_sheet[8]:
        cell.font = bold
    for line in lines:
        quote_sheet.append([line["label"], line["description"], round_money(line["amount"])])
    quote_sheet.append([])
    quote_sheet.append(["Subtotal (ex VAT)", round_money(header["subtotal_ex_vat"])])
    quote_sheet.append(["VAT", round_money(header["vat_amount"])])
    quote_sheet.append(["Total (inc VAT)", round_money(header["total_inc_vat"])])
    quote_sheet.append([])
    quote_sheet.append(["Payment terms", header["payment_terms"]])

    internal = wb.create_sheet("Internal")
    internal["A1"] = "Internal commercial summary"
    internal["A1"].font = bold
    internal.append(["Reference", data.reference])
    internal.append(["Customer", data.customer_name])
    internal.append(["Materials cost", data.materials_cost])
    internal.append(["Labour cost", data.labour_cost])
    internal.append(["Waste cost", data.waste_cost])
    internal.append(["Travel cost", data.travel_cost])
    internal.append(["Prelim cost", data.prelim_cost])
    internal.append(["Total cost", data.total_cost])
    internal.append(["Calculated sell", data.calculated_sell_price])
    internal.append(["Final sell", data.sell_price])
    internal.append(["Margin £", data.margin_value])
    internal.append(["Margin %", data.margin_percent])
    internal.append(["Min job applied", "Yes" if data.min_job_applied else "No"])
    internal.append([])
    header_row = internal.max_row + 1
    internal.append(["Work type", "Line cost", "Line sell", "Target margin %"])
    for cell in internal[header_row]:
        cell.font = bold
    for item in data.items:
        internal.append(
            [
                item.label,
                round_money(item.line_cost),
                round_money(item.line_sell),
                item.target_margin_percent,
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    filename = _safe_filename(header["reference"], header["customer_name"], "Estimate.xlsx")
    return out.getvalue(), filename


def render_estimates_list_csv(estimates: Iterable[Estimate]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "Reference",
            "Customer",
            "Site",
            "Postcode",
            "Status",
            "Sell (ex VAT)",
            "Margin %",
            "Surveyor",
            "Survey date",
        ]
    )
    for row in estimates:
        writer.writerow(
            [
                row.reference,
                row.customer_name,
                row.site_address,
                row.postcode,
                row.status,
                f"{row.sell_price:.2f}",
                f"{row.margin_percent:.2f}",
                row.surveyor,
                row.survey_date,
            ]
        )
    return buffer.getvalue().encode("utf-8-sig")
